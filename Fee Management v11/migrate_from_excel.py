"""
Migration script: Imports all data from Account_Management_2026-2027_v1.xlsm
Run once: python migrate_from_excel.py
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from database.db import init_db, get_db
from datetime import datetime

EXCEL_PATH = "/mnt/user-data/uploads/Account_Management_2026-2027_v1.xlsm"

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

def safe_float(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

def safe_int(v):
    try: return int(v) if v is not None else 0
    except: return 0

def fmt_date(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%d-%m-%Y')
    return safe_str(v)

def fmt_ts(v):
    if v is None: return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d %H:%M:%S')
    return safe_str(v)

def run():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True, data_only=True)

    init_db()
    conn = get_db()
    c = conn.cursor()

    # ── CLEAR EXISTING (fresh migration) ────────────────────────────────
    for tbl in ['students','counters','used_prefixes','receipt_log','misc_transactions']:
        c.execute(f"DELETE FROM {tbl}")
    # Re-seed counters
    for name, val in [("new_admission",0),("re_admission",0),
                       ("tuition_receipt",0),("misc_receipt",0)]:
        c.execute("INSERT OR IGNORE INTO counters (name,value) VALUES (?,?)", (name,val))

    # ── 1. STUDENTS ─────────────────────────────────────────────────────
    ws = wb['Database']
    inserted = 0
    skipped = 0
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=20, values_only=True):
        # col mapping:
        # 0=date_adm, 1=sl_no, 2=student_name, 3=class, 4=section, 5=ref_no
        # 6=father, 7=mother, 8=father_contact, 9=mother_contact, 10=status
        # 11=gross, 12=concession, 13=net, 14=neft, 15=cash, 16=balance, 17=comments
        name = safe_str(row[2])
        if not name or name == 'None':
            skipped += 1
            continue
        # Skip the totals row
        if row[1] is None and row[11] is not None and name == '':
            continue

        ref = safe_str(row[5]) or None
        gross = safe_float(row[11])
        conc  = safe_float(row[12])
        net   = safe_float(row[13])
        neft  = safe_float(row[14])
        cash  = safe_float(row[15])
        bal   = safe_float(row[16])

        try:
            c.execute("""INSERT OR IGNORE INTO students
                (date_of_admission, sl_no, student_name, class, section,
                 reference_no, father_name, mother_name,
                 father_contact, mother_contact, status,
                 gross_tuition, concession, net_tuition,
                 neft_paid, cash_paid, balance, comments)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    fmt_date(row[0]),
                    safe_int(row[1]),
                    name,
                    safe_str(row[3]),
                    safe_str(row[4]),
                    ref,
                    safe_str(row[6]),
                    safe_str(row[7]),
                    safe_str(row[8]),
                    safe_str(row[9]),
                    safe_str(row[10]) or 'Not Admitted',
                    gross, conc, net, neft, cash, bal,
                    safe_str(row[17]),
                ))
            inserted += 1
        except Exception as e:
            print(f"  Skipping student '{name}': {e}")
            skipped += 1

    print(f"  Students: {inserted} inserted, {skipped} skipped")

    # ── 2. COUNTERS ──────────────────────────────────────────────────────
    ws_ctrl = wb['Control']
    ctrl_map = {
        'New Admission': 'new_admission',
        'Re-Admission':  're_admission',
        'Tuition Receipt Counter': 'tuition_receipt',
        'MISC. Receipt Counter':   'misc_receipt',
    }
    for row in ws_ctrl.iter_rows(min_row=2, max_row=10, values_only=True):
        if row[0] in ctrl_map:
            c.execute("UPDATE counters SET value=? WHERE name=?",
                      (safe_int(row[1]), ctrl_map[row[0]]))
            print(f"  Counter '{ctrl_map[row[0]]}' = {row[1]}")

    # ── 3. USED PREFIXES ─────────────────────────────────────────────────
    ws_pfx = wb['USED_PREFIXES']
    pfx_count = 0
    for row in ws_pfx.iter_rows(min_row=2, values_only=True):
        p = safe_str(row[0])
        if p:
            c.execute("INSERT OR IGNORE INTO used_prefixes (prefix) VALUES (?)", (p,))
            pfx_count += 1
    print(f"  Used prefixes: {pfx_count}")

    # ── 4. RECEIPT LOG (GOOGLE_QUEUE) ─────────────────────────────────────
    ws_gq = wb['GOOGLE_QUEUE']
    rcpt_count = 0
    for row in ws_gq.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue  # no receipt no
        try:
            c.execute("""INSERT OR IGNORE INTO receipt_log
                (timestamp, receipt_no, student_name, amount, payment_mode, fee_type, created_by, status)
                VALUES (?,?,?,?,?,?,'system','active')""",
                (fmt_ts(row[0]), safe_str(row[1]), safe_str(row[2]),
                 safe_float(row[3]), safe_str(row[4]), safe_str(row[5])))
            rcpt_count += 1
        except Exception as e:
            print(f"  Receipt log skip: {e}")
    print(f"  Receipt log entries: {rcpt_count}")

    # ── 5. MISC TRANSACTIONS ─────────────────────────────────────────────
    ws_misc = wb['MISC.']
    misc_count = 0
    for row in ws_misc.iter_rows(min_row=2, max_col=7, values_only=True):
        if not row[1]: continue  # no student name
        try:
            c.execute("""INSERT INTO misc_transactions
                (timestamp, student_name, class, reference_no, payment_mode, amount, fee_type, status)
                VALUES (?,?,?,?,?,?,?,'active')""",
                (fmt_ts(row[0]), safe_str(row[1]), safe_str(row[2]),
                 safe_str(row[3]), safe_str(row[4]),
                 safe_float(row[5]), safe_str(row[6]) or 'MISC'))
            misc_count += 1
        except Exception as e:
            print(f"  MISC skip: {e}")
    print(f"  MISC transactions: {misc_count}")

    conn.commit()
    conn.close()
    print("\n✅ Migration complete!")

    # Summary
    conn2 = get_db()
    total = conn2.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    na    = conn2.execute("SELECT COUNT(*) FROM students WHERE reference_no LIKE 'NA%'").fetchone()[0]
    ra    = conn2.execute("SELECT COUNT(*) FROM students WHERE reference_no LIKE 'RA%'").fetchone()[0]
    wait  = conn2.execute("SELECT COUNT(*) FROM students WHERE reference_no IS NULL OR reference_no=''").fetchone()[0]
    print(f"\n   Total students : {total}")
    print(f"   NA students    : {na}")
    print(f"   RA students    : {ra}")
    print(f"   Waiting (no ref): {wait}")
    conn2.close()

if __name__ == "__main__":
    run()
