"""School Fee Management System v3 - Network Ready"""
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import os, sys, socket

# ── PyInstaller / EXE resource path support ──────────────────────────────────
def resource_path(relative_path):
    """Works both in normal Python and when bundled as .exe by PyInstaller"""
    try:
        base_path = sys._MEIPASS  # PyInstaller extracts here
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def data_path(relative_path):
    """For writable data (database, receipts) — always next to the exe/script"""
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)

sys.path.insert(0, data_path('.'))
from database.db import init_db, get_db, set_db_path
from modules.students import (search_students, add_student, get_all_students, get_students_no_ref,
                               delete_student, get_student_by_ref, update_student_field,
                               update_tuition_fees, update_phone, assign_ref_to_existing)
from modules.pdf_generator import set_receipts_path
from modules.receipts import (generate_receipt, get_receipt_log, get_receipt_counts,
                               cancel_receipt, search_receipts, get_cancelled_receipts)
from modules.misc import get_misc_summary, get_misc_transactions, update_misc_to_pay, nuke_all_misc
from modules.dashboard import get_dashboard_stats
from modules.daily_report import get_daily_report_data, log_daily_report, get_report_log, delete_old_reports
from modules.utils import rupees_in_words
from modules.auth import (login_user, verify_password, get_all_users, create_user,
                           update_password, toggle_user, audit, get_audit_log, rename_user)

app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static")
)
app.secret_key = "school_fee_v3_2026_secure"
# Sessions expire after 8 hours of inactivity (school day)
app.config['PERMANENT_SESSION_LIFETIME'] = __import__('datetime').timedelta(hours=8)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

@app.before_request
def make_session_permanent():
    session.permanent = True

# Set writable paths for database and receipts
set_db_path(data_path(os.path.join("database", "school_fees.db")))

# Academic year folder name: Receipts_YYYY-YY (e.g. Receipts_2026-27)
import platform as _platform
from datetime import datetime as _dt, date
def _academic_year():
    y = _dt.now().year
    m = _dt.now().month
    # Academic year: March onwards = current year start (e.g. Mar 2026 = 2026-27)
    if m >= 3:
        return f"{y}-{str(y+1)[2:]}"
    else:
        return f"{y-1}-{str(y)[2:]}"

_app_dir = os.path.dirname(os.path.abspath(__file__))
_folder_name = f"Receipts_{_academic_year()}"
if _platform.system() == "Windows":
    _receipts_dir = os.path.join(os.path.expanduser("~"), "Desktop", _folder_name)
else:
    _receipts_dir = os.path.join(os.path.dirname(_app_dir), _folder_name)
os.makedirs(os.path.join(_receipts_dir, "TUITION"),   exist_ok=True)
os.makedirs(os.path.join(_receipts_dir, "MISC"),      exist_ok=True)
os.makedirs(os.path.join(_receipts_dir, "CANCELLED"), exist_ok=True)
set_receipts_path(_receipts_dir)

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def logged_in(): return session.get("user_id") is not None
def current_user(): return session.get("username","system")
def current_role(): return session.get("role","")

def require_login(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**k):
        if not logged_in():
            # API routes must return JSON, not redirect to login HTML
            from flask import request as _req
            if _req.path.startswith('/api/'):
                return jsonify({"success":False,"error":"Session expired. Please log in again.","logged_in":False}), 401
            return redirect(url_for("login_page"))
        return f(*a,**k)
    return d

@app.before_request
def startup(): init_db()

# ── AUTH ─────────────────────────────────────────────────────────────────────
@app.route("/login")
def login_page():
    if logged_in(): return redirect("/")
    return render_template("login.html")

@app.route("/api/login", methods=["POST"])
def api_login():
    d = request.json
    result = login_user(d.get("username",""), d.get("password",""))
    if result["success"]:
        u = result["user"]
        session.update({"user_id":u["id"],"username":u["username"],"role":u["role"],"display_name":u["display_name"] or u["username"]})
        audit(u["username"],u["role"],"LOGIN","Logged in",request.remote_addr)
    return jsonify(result)

@app.route("/logout")
def logout():
    if logged_in(): audit(current_user(),current_role(),"LOGOUT","Logged out")
    session.clear()
    return redirect("/login")

@app.route("/api/session")
def api_session():
    return jsonify({"logged_in":logged_in(),"username":current_user(),"role":current_role(),"display_name":session.get("display_name","") or current_user()})

# ── PAGES ─────────────────────────────────────────────────────────────────────
@app.route("/")
@require_login
def index(): return redirect("/receipt")

@app.route("/receipt")
@require_login
def receipt_generator(): return render_template("receipt_generator.html")

@app.route("/new-admission")
@require_login
def page_new_admission(): return render_template("new_admission.html")

@app.route("/re-admission")
@require_login
def page_re_admission(): return render_template("re_admission.html")

@app.route("/students")
@require_login
def students(): return render_template("students.html")

@app.route("/total-balance")
@require_login
def total_balance(): return render_template("total_balance.html")

@app.route("/misc-transactions")
@require_login
def misc_transactions_page(): return render_template("misc_transactions.html")

@app.route("/all-receipts")
@require_login
def all_receipts(): return render_template("all_receipts.html")

@app.route("/dashboard")
@require_login
def dashboard(): return render_template("dashboard.html")

@app.route("/tc")
@require_login
def tc_page(): return render_template("tc.html")

@app.route("/cancel-receipt")
@require_login
def cancel_receipt_page(): return render_template("cancel_receipt.html")

@app.route("/daily-report")
@require_login
def daily_report(): return render_template("daily_report.html")

@app.route("/management")
@require_login
def management():
    if current_role()!="management": return redirect("/")
    return render_template("management.html")

# ── API: STUDENTS ─────────────────────────────────────────────────────────────
@app.route("/api/students/search")
@require_login
def api_search_students():
    no_ref = request.args.get("no_ref","0")=="1"
    return jsonify(search_students(request.args.get("q",""), request.args.get("field","student"), no_ref))

@app.route("/api/students/no-ref")
@require_login
def api_students_no_ref():
    return jsonify(get_students_no_ref())

@app.route("/api/students/all")
@require_login
def api_all_students(): return jsonify(get_all_students())

@app.route("/api/students/add", methods=["POST"])
@require_login
def api_add_student():
    data = request.json
    adm_type = data.pop("admission_type","NA")
    if adm_type == "RA":
        from modules.admissions import re_admission as do_re
        return jsonify(do_re(data, current_user()))
    else:
        from modules.admissions import new_admission as do_na
        return jsonify(do_na(data, current_user()))

@app.route("/api/students/assign-ref", methods=["POST"])
@require_login
def api_assign_ref():
    d = request.json
    return jsonify(assign_ref_to_existing(d["student_id"], d["admission_type"], d, current_user()))

@app.route("/api/students/delete", methods=["POST"])
@require_login
def api_delete_student():
    d = request.json
    return jsonify(delete_student(d.get("student_name"), d.get("reason","TC"), current_user()))

@app.route("/api/students/update-field", methods=["POST"])
@require_login
def api_update_field():
    d = request.json
    return jsonify(update_student_field(d["student_id"], d["field"], d["value"], current_user()))

@app.route("/api/students/update-tuition", methods=["POST"])
@require_login
def api_update_tuition():
    d = request.json
    return jsonify(update_tuition_fees(d["student_id"], d["gross"], d["concession"], current_user()))

@app.route("/api/students/update-phone", methods=["POST"])
@require_login
def api_update_phone():
    d = request.json
    return jsonify(update_phone(d["student_id"], d.get("father_contact",""), d.get("mother_contact",""), current_user()))

@app.route("/api/students/by-ref/<ref_no>")
@require_login
def api_student_by_ref(ref_no): return jsonify(get_student_by_ref(ref_no))

@app.route("/api/students/balance")
@require_login
def api_student_balance():
    """Returns tuition balance and misc balance for a student"""
    name = request.args.get("name","")
    conn = get_db()
    try:
        s = conn.execute(
            "SELECT net_tuition, neft_paid, cash_paid, misc_to_pay FROM students WHERE student_name=?",
            (name,)
        ).fetchone()
        if not s:
            return jsonify({"found": False})
        net      = s["net_tuition"] or 0
        neft     = s["neft_paid"] or 0
        cash     = s["cash_paid"] or 0
        misc_pay = s["misc_to_pay"] or 0
        tuition_balance = net - neft - cash
        misc_paid = conn.execute(
            """SELECT COALESCE(SUM(amount),0) FROM misc_transactions
               WHERE student_name=? AND status='active'""", (name,)
        ).fetchone()[0]
        misc_balance = misc_pay - misc_paid
        return jsonify({
            "found": True,
            "tuition_balance": tuition_balance,
            "misc_balance": misc_balance,
            "net_tuition": net,
            "neft_paid": neft,
            "cash_paid": cash,
            "misc_to_pay": misc_pay,
            "misc_paid": misc_paid,
        })
    finally:
        conn.close()

# ── API: RECEIPTS ─────────────────────────────────────────────────────────────
@app.route("/api/receipt/generate", methods=["POST"])
@require_login
def api_generate_receipt(): return jsonify(generate_receipt(request.json, current_user()))

@app.route("/api/receipt/download/<path:receipt_no>")
@require_login
def api_download_receipt(receipt_no):
    from modules.pdf_generator import get_receipts_base
    # Search the correct receipts folder (Desktop/Receipts_YYYY-YY on Windows)
    receipts_dir = get_receipts_base()
    for root, dirs, files in os.walk(receipts_dir):
        for f in files:
            if receipt_no in f and f.endswith(".pdf"):
                return send_file(os.path.join(root, f), as_attachment=True, download_name=f)
    # PDF missing from disk — regenerate it on the fly from DB
    conn = get_db()
    try:
        r = conn.execute(
            "SELECT * FROM receipt_log WHERE receipt_no=? AND status='active'",
            (receipt_no,)).fetchone()
        if r:
            from modules.pdf_generator import generate_receipt_pdf
            from modules.utils import rupees_in_words
            result = generate_receipt_pdf({
                "student_name":  r["student_name"] or "",
                "parent_name":   r["parent_name"]  or "",
                "grade":         r["grade"]         or "",
                "section":       r["section"]       or "",
                "reference_no":  r["reference_no"]  or "",
                "receipt_no":    r["receipt_no"],
                "payment_mode":  r["payment_mode"],
                "payment_date":  r["payment_date"],
                "fee_type":      r["fee_type"],
                "amount":        r["amount"],
                "amount_words":  rupees_in_words(r["amount"]),
            }, r["fee_type"])
            if result.get("pdf_path") and os.path.exists(result["pdf_path"]):
                conn.execute("UPDATE receipt_log SET pdf_path=? WHERE receipt_no=?",
                             (result["pdf_path"], receipt_no))
                conn.commit()
                return send_file(result["pdf_path"], as_attachment=True,
                                 download_name=result["pdf_filename"])
    finally:
        conn.close()
    return jsonify({"error": "PDF not found — try regenerating from the receipt page"}), 404

@app.route("/api/receipt/log")
@require_login
def api_receipt_log(): return jsonify(get_receipt_log(int(request.args.get("page",1)), 200, request.args.get("fee_type","")))

@app.route("/api/receipt/all")
@require_login
def api_receipt_all():
    conn = get_db()
    try:
        rows = conn.execute("""SELECT receipt_no,student_name,parent_name,grade,section,
                                      reference_no,amount,payment_mode,payment_date,fee_type,
                                      pdf_path,created_by,status,created_at
                               FROM receipt_log ORDER BY timestamp DESC""").fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()

@app.route("/api/receipt/counts")
@require_login
def api_receipt_counts(): return jsonify(get_receipt_counts())

@app.route("/api/receipt/search")
@require_login
def api_search_receipts(): return jsonify(search_receipts(request.args.get("q","")))

@app.route("/api/receipt/cancel", methods=["POST"])
@require_login
def api_cancel_receipt():
    d = request.json
    if not verify_password(current_user(), d.get("password","")):
        return jsonify({"success":False,"error":"Incorrect password. Cancellation denied."})
    return jsonify(cancel_receipt(d.get("receipt_no"), d.get("reason"), current_user()))

@app.route("/api/receipt/cancelled")
@require_login
def api_cancelled_receipts(): return jsonify(get_cancelled_receipts())

# ── API: MISC ─────────────────────────────────────────────────────────────────
@app.route("/api/misc/summary")
@require_login
def api_misc_summary(): return jsonify(get_misc_summary())

@app.route("/api/misc/transactions")
@require_login
def api_misc_transactions(): return jsonify(get_misc_transactions())

@app.route("/api/misc/update-to-pay", methods=["POST"])
@require_login
def api_update_misc_to_pay():
    d = request.json
    return jsonify(update_misc_to_pay(d["reference_no"], float(d["amount"]), current_user()))

# ── API: DASHBOARD ───────────────────────────────────────────────────────────
@app.route("/api/dashboard/stats")
@require_login
def api_dashboard_stats(): return jsonify(get_dashboard_stats())

# ── API: DAILY REPORT ─────────────────────────────────────────────────────────
@app.route("/api/daily-report")
@require_login
def api_daily_report():
    pw = request.args.get("password","")
    if not pw or not verify_password(current_user(), pw):
        return jsonify({"success":False,"error":"Password required to generate daily report."})
    report_date = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    data = get_daily_report_data(report_date)
    if data.get("success"):
        log_daily_report(report_date, current_user(), data)
        delete_old_reports(60)
    return jsonify(data)

@app.route("/api/daily-report/log")
@require_login
def api_daily_report_log():
    return jsonify(get_report_log())

# ── API: UTILS ───────────────────────────────────────────────────────────────
@app.route("/api/utils/rupees-in-words")
def api_rupees_in_words():
    try: return jsonify({"words":rupees_in_words(float(request.args.get("amount",0)))})
    except: return jsonify({"words":""})

# ── API: MANAGEMENT ──────────────────────────────────────────────────────────
@app.route("/api/management/users")
@require_login
def api_get_users():
    if current_role()!="management": return jsonify([])
    return jsonify(get_all_users())

@app.route("/api/management/create-user", methods=["POST"])
@require_login
def api_create_user():
    if current_role()!="management": return jsonify({"success":False,"error":"Access denied"})
    d=request.json
    return jsonify(create_user(d["username"],d["password"],d["role"],d.get("display_name",""),current_user()))

@app.route("/api/management/change-password", methods=["POST"])
@require_login
def api_change_password():
    if current_role()!="management": return jsonify({"success":False,"error":"Access denied"})
    d=request.json
    return jsonify(update_password(d["username"],d["new_password"],current_user()))

@app.route("/api/management/toggle-user", methods=["POST"])
@require_login
def api_toggle_user():
    if current_role()!="management": return jsonify({"success":False,"error":"Access denied"})
    d=request.json
    return jsonify(toggle_user(d["username"],d["active"],current_user()))

@app.route("/api/management/rename-user", methods=["POST"])
@require_login
def api_rename_user():
    if current_role()!="management": return jsonify({"success":False,"error":"Access denied"})
    d=request.json
    return jsonify(rename_user(d["old_username"],d["new_username"],d.get("display_name",""),current_user()))

@app.route("/api/management/audit-log")
@require_login
def api_audit_log():
    if current_role()!="management": return jsonify({"logs":[],"total":0})
    return jsonify(get_audit_log(
        int(request.args.get("page",1)),
        200,
        request.args.get("username",""),
        request.args.get("action","")
    ))

# ── EXCEL EXPORT ─────────────────────────────────────────────────────────────
@app.route("/api/students/export-excel")
@require_login
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    students = get_all_students()
    wb = Workbook(); ws = wb.active; ws.title = "Student Database"
    # Header style - blue
    hf = PatternFill("solid", fgColor="1A5276")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    # ALL columns - every field regardless of screen visibility
    headers = [
        "#", "Student UID", "Ref No.", "Date of Admission", "Sl No.",
        "Student Name", "RTE", "Father Name", "Mother Name",
        "Father Contact", "Mother Contact", "Class", "Section", "Status",
        "Gross Tuition", "Concession", "Net Tuition", "MISC To Pay",
        "NEFT Paid", "Cash Paid", "Balance (Tuition)", "Comments"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    # Data rows
    for i, s in enumerate(students, 1):
        balance = (s["net_tuition"] or 0) - (s["neft_paid"] or 0) - (s["cash_paid"] or 0)
        fill = PatternFill("solid", fgColor="EBF5FB") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row = [
            i,
            s.get("student_uid") or "",
            s["reference_no"] or "",
            s["date_of_admission"] or "",
            s["sl_no"] or "",
            s["student_name"] or "",
            "Yes" if s.get("is_rte") else "",
            s["father_name"] or "",
            s["mother_name"] or "",
            s["father_contact"] or "",
            s["mother_contact"] or "",
            s["class"] or "",
            s["section"] or "",
            s["status"] or "",
            s["gross_tuition"] or 0,
            s["concession"] or 0,
            s["net_tuition"] or 0,
            s["misc_to_pay"] or 0,
            s["neft_paid"] or 0,
            s["cash_paid"] or 0,
            balance,
            s["comments"] or ""
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(i + 1, col).fill = fill
    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)
    # Freeze header row
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Student_Database.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/misc/export-excel")
@require_login
def api_export_misc_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    summary = get_misc_summary()
    wb = Workbook(); ws = wb.active; ws.title = "Total Balance"
    hf = PatternFill("solid", fgColor="1E8449")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    # ALL columns including UID and RTE
    headers = [
        "#", "Student UID", "Ref No.", "Student Name", "Class", "Section", "RTE",
        "Gross Tuition", "Concession", "Net Tuition", "MISC To Pay",
        "Total Payable", "Tuition Paid (NEFT+Cash)", "MISC Paid", "Total Paid", "Balance"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for i, r in enumerate(summary, 1):
        fill = PatternFill("solid", fgColor="EAFAF1") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        # Get gross/concession from the summary or fall back to 0
        gross = r.get("gross_tuition", 0) or 0
        conc  = r.get("concession", 0) or 0
        row = [
            i,
            r.get("student_uid") or "",
            r["reference_no"] or "",
            r["student_name"] or "",
            r["class"] or "",
            r["section"] or "",
            "Yes" if r.get("is_rte") else "",
            gross,
            conc,
            r["net_tuition"] or 0,
            r["misc_to_pay"] or 0,
            r["total_payable"] or 0,
            r["tuition_paid"] or 0,
            r["misc_paid"] or 0,
            r["total_paid"] or 0,
            r["balance"] or 0,
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(i + 1, col).fill = fill
    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Total_Balance.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/nuke-misc", methods=["POST"])
@require_login
def api_nuke_misc():
    d = request.json
    return jsonify(nuke_all_misc(d.get("password", "")))


if __name__ == "__main__":
    ip = get_local_ip()
    init_db()

    # Check if already running
    import socket as _sock
    with _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM) as _s:
        if _s.connect_ex(('localhost', 5000)) == 0:
            print("\n" + "="*56)
            print("  App already running! Opening browser...")
            print("="*56)
            import webbrowser
            webbrowser.open("http://localhost:5000")
            input("Press Enter to close...")
            import sys; sys.exit(0)

    print("\n" + "="*56)
    print("  SCHOOL FEE MANAGEMENT SYSTEM v3")
    print("="*56)
    print(f"  This computer : http://localhost:5000")
    print(f"  Other devices : http://{ip}:5000")
    print(f"  Share second link with other school computers")
    print("-"*56)
    print("  management / Mgmt@2026")
    print("  admin1     / Admin1@2026")
    print("  admin2     / Admin2@2026")
    print("-"*56)
    print("  KEEP THIS WINDOW OPEN while using the app.")
    print("="*56 + "\n")

    import threading, webbrowser, time
    def open_browser():
        time.sleep(2)
        webbrowser.open("http://localhost:5000")
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(debug=False, host="0.0.0.0", port=5000, use_reloader=False)

